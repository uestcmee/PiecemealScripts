import os
import openpyxl
from copy import copy


dir='C:\\Users\\zikep\\Desktop\\20200206发改委\\关于做好全国优化营商环境先进经验和典型做法复制\\昆明\\昆明经验总结'
file_names = os.listdir(dir)
for filename in file_names:
    if filename[-4:]=='.xls':
        print(filename)
        continue
    print(filename)
    wb=openpyxl.load_workbook(filename=dir+'\\'+filename)
    ws=wb['Sheet1']
    old_title=ws['A2'].value

    new_title=old_title.replace('太原','昆明')
    print(new_title)
    print(ws.max_row)

    #删除第四行之后的部分
    ws.delete_rows(4,ws.max_row-3)


    #修改title
    ws['A2']=new_title
    wb.save(dir+'\\'+filename)
